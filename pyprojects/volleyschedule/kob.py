import sys
sys.path.append("../")
import numpy as np
import itertools
import random
import json
import myhelp


class KoB:

    def __init__(self, num_pairs, num_courts, num_games_per_pair, maxbyes_allowed, maxrounds_allowed):
        self._num_pairs = num_pairs
        self._num_courts = num_courts
        self._num_games_per_pair = num_games_per_pair
        self._maxbyes_allowed = maxbyes_allowed
        self._maxrounds_allowed = maxrounds_allowed



    def checkValid(self, game, valid_games, count_games=None, strict=0):
        game_set = set(game)
        is_valid = True
        if len(game_set) != 4:
            is_valid = False
        for valid_game in valid_games:
            valid_game_set = set(valid_game)
            check_set = valid_game_set.union( game_set )
            if strict==0:
                if len(check_set) <= 6:
                    is_valid = False
            elif strict==1:
                opponent = [ set([valid_game[0],valid_game[1]]), \
                             set([valid_game[2],valid_game[3]]), \
                             set([game[0],game[1]]),    \
                             set([game[2],game[3]]),    \
                           ]
                if     len(opponent[0].union(opponent[2])) <= 2 \
                    or len(opponent[0].union(opponent[3])) <= 2 \
                    or len(opponent[1].union(opponent[2])) <= 2 \
                    or len(opponent[1].union(opponent[3])) <= 2 :
                    is_valid = False

        if count_games is not None:                
            for n in game:
                if count_games[n] >= self._num_games_per_pair:
                    is_valid = False

        return is_valid



    def addGame(self, game, valid_games, count_games):
        idxmin = np.argmin( count_games )
        if idxmin in game:
            for n in game:
                count_games[n] += 1
            valid_games.append( game )
            return True

        return False



    def formGamesWithPolicy(self, num_segs, num_choices_per_seg, order=0):
        num_segs_to_choose_from = 4 // num_choices_per_seg
        seg_size = self._num_pairs // num_segs
        games = []

        # inter-segment
        for seg_index in itertools.combinations(range(num_segs), num_segs_to_choose_from):
            for inseg_index in itertools.product(range(seg_size), repeat=4):
                game = [ seg_index[n//num_choices_per_seg]*seg_size + inseg_index[n] for n in range(4) ]
                if order == 2:
                    def f(i):
                        if i % 2 == 0:
                            return i
                        else:
                            return self._num_pairs - i
                    game = [ f(i)  for i in game ]

                elif order == 1:
                    game = [ self._num_pairs-1-i for i in game ]
                
                gamelist = [ 
                            [game[0], game[1], game[2], game[3]],
                            [game[0], game[2], game[1], game[3]],
                            [game[0], game[3], game[1], game[2]],
                            ]
                games += gamelist

        return games



    def makeGamesByRound(self):
        self._games = []
        

    def makeGames(self):
        self._games = []
        count_games = np.zeros( self._num_pairs, dtype=int )
        
        args_list = [ 
                      [4, 1, 0, 0],
                      [4, 1, 2, 1],
                    ]

        for args in args_list:
            if_any_added = True
            while if_any_added:
                if_added = False
                games = self.formGamesWithPolicy(args[0], args[1], order=args[2])
                for game in games:
                    if_valid = self.checkValid(game, self._games, count_games=count_games, strict=args[3])
                    if if_valid:
                        if_added = if_added or self.addGame(game, self._games, count_games)
                if_any_added = if_added

        return self._games
            

    
    def checkGames(self):
        print( "Game-making results:" )
        for g in self._games:
            print( f"{g[0]:2} {g[1]:2} vs {g[2]:2} {g[3]:2}" )
        print( f"A total number of {len(self._games)} games." )
        for n in range(self._num_pairs):
            games = []
            teammates = []
            opponents = []
            for g in self._games:
                if n in g:
                    games.append( g )
                    if n in g[0:2]:
                        teammates += [i for i in g[0:2] if i != n]
                        opponents += g[2:4]
                    elif n in g[2:4]:
                        teammates += [i for i in g[2:4] if i != n] 
                        opponents += g[0:2] 
            opponents.sort()
            teammates.sort()
            mates = opponents + teammates
            print( f"pair {n:2} plays {len(games):2} games with/against m/m/m {min(mates):4} {np.median(mates):6} {max(mates):4}, encounters {len(set(mates))} pairs" )
            if len(set(teammates)) < self._num_games_per_pair:
                print( f"teammates are {teammates} and opponents are {opponents}" )


    def countByes(self, n, games):
        game_rank = []
        for i,g in enumerate(games):
            if n in g:
                game_rank.append( i )
        game_round = [ r//self._num_courts for r in game_rank]
        game_round.sort()
        if len(game_round) <= 1:
            minbyes = 0
            maxbyes = 0
        else:
            minbyes = 9999
            maxbyes = 0
            for r in range(1, len(game_round)):
                byes = game_round[r] - game_round[r-1] - 1
                if byes < minbyes:
                    minbyes = byes
                if byes > maxbyes:
                    maxbyes = byes

        return game_round, minbyes, maxbyes



    def checkGameSchedule(self, all_games):
        print( "Game schdule results:" )
        for game in all_games:
            if len(game) > 0:
                print( f"{game[0]:2} {game[1]:2} vs {game[2]:2} {game[3]:2}" )
        print( f"A total number of {len(all_games)} games." )

        for pair in range(self._num_pairs):
            byes = self.countByes(pair, all_games)
            playsit = []
            j = 0
            for i in range(self._maxrounds_allowed):
                if j < self._num_games_per_pair and i == byes[0][j]:
                    playsit.append( "play" )
                    j += 1
                else:
                    playsit.append( "sit" )

            games = []
            teammates = []
            opponents = []
            for game in all_games:
                if pair in game:
                    games.append( game )
                    if pair in game[0:2]:
                        teammates += [i for i in game[0:2] if i != pair]
                        opponents += game[2:4]
                    elif pair in game[2:4]:
                        teammates += [i for i in game[2:4] if i != pair] 
                        opponents += game[0:2] 
            opponents.sort()
            teammates.sort()
            mates = opponents + teammates
            opponents_no_reps = list(set( opponents ))

            s = "Pair "
            s += (pair+1).__str__().rjust(2)
            s += " plays "
            s += len(games).__str__().rjust(2)
            s += " games, with "
            s += "/".join( ((i+1).__str__().rjust(2) for i in teammates) )
            s += ", against "
            s += "/".join( ((i+1).__str__().rjust(2) for i in opponents_no_reps) )
            s += " ("
            s += (len(opponents)-len(opponents_no_reps)).__str__()
            s += " repeats)."
            s += " Follows "
            s += "->".join( playsit )
            s += " ("
            s += byes[2].__str__()
            s += " consecutive byes max)."

            print( s )



    def scheduleGames(self, *args, **kwargs):
        games_left = self._games.copy()
        random.shuffle( games_left )
        games_scheduled = []

        while len(games_left) > 0:
            round_scheduled_list = []
            index_game_list = []
            for index_game,game in enumerate(games_left):
                round_scheduled = -1
                for n in game:
                    for index_game_scheduled,game_scheduled in enumerate(games_scheduled):
                        if n in game_scheduled:
                            round_scheduled = max(round_scheduled, index_game_scheduled // self._num_courts)

                round_scheduled_list.append( round_scheduled )
                index_game_list.append( index_game )

            round_scheduled_list, index_game_list = zip(*sorted(zip(round_scheduled_list, index_game_list)))
               
            round_now = len(games_scheduled)//self._num_courts
            index_game_to_add = index_game_list[0]

            if round_scheduled_list[0] < round_now:
                games_scheduled.append( games_left[index_game_to_add] )
                games_left.pop( index_game_to_add )
            else:
                games_scheduled.append( [] )

        maxbyes = 0
        for n in range(self._num_pairs):    
            byes = self.countByes(n, games_scheduled)
            if byes[-1] > maxbyes:
                maxbyes = byes[-1]
        num_rounds = np.ceil( len(games_scheduled)/self._num_courts )

        if maxbyes<=2:
            for n in range(self._num_pairs):    
                byes = self.countByes(n, games_scheduled)
                print(n, " -> ", byes )
 
        return maxbyes, num_rounds, games_scheduled



    def scheduleGamesB(self, *args, **kwargs):
        games_left = self._games.copy()
        random.shuffle( games_left )
        games_scheduled = []
        last_round_rank = {n: -1 for n in range(self._num_pairs)}

        while len(games_left) > 0:
            index_games_scheduled_start_current_round = len(games_scheduled) // self._num_courts * self._num_courts
            index_games_ok_to_add = []
            for index_game,game in enumerate(games_left):
                is_ok_to_add = True
                for n in game:
                    for game_scheduled in games_scheduled[index_games_scheduled_start_current_round:]:
                        if n in game_scheduled:
                            is_ok_to_add = False
                if is_ok_to_add:
                    index_games_ok_to_add.append( index_game )
           
            if len(index_games_ok_to_add) > 0: 
                min_last_round_rank = [ min(last_round_rank[n] for n in games_left[i]) for i in index_games_ok_to_add  ]
                minrank, minindex = zip(*sorted(zip(min_last_round_rank, index_games_ok_to_add)))
#                print( "MR", minrank )
#                print( "MI", minindex )
                minindex = [ minindex[i] for i in range(len(minindex)) if minrank[i]==minrank[0] ]
                max_last_round_rank = [ max(last_round_rank[n] for n in games_left[i]) for i in minindex ]
                rank, index = zip(*sorted(zip(max_last_round_rank, minindex), reverse=True))
#                print( "R", rank )
#                print( "I", index )
                
                games_scheduled.append( games_left[index[0]] )
                for n in games_left[index[0]]:
                    last_round_rank[n] = (len(games_scheduled)-1) // self._num_courts
                games_left.pop( index[0] )
            else:
                games_scheduled.append( [] )

#            print( "Schedule: " )
#            for n in range(0, len(games_scheduled), 5):
#                print(games_scheduled[n:min(n+5,len(games_scheduled))] )

        maxbyes = 0
        for n in range(self._num_pairs):    
            byes = self.countByes(n, games_scheduled)
            if byes[-1] > maxbyes:
                maxbyes = byes[-1]
        num_rounds = np.ceil( len(games_scheduled)/self._num_courts )

        if True or maxbyes<=2:
            print("num_rounds = ", num_rounds)
            for n in range(self._num_pairs):    
                byes = self.countByes(n, games_scheduled)
                print(n, " -> ", byes )

        return maxbyes, num_rounds, games_scheduled


    def saveGameSchedule(self):
        maxbyes = 999
        num_rounds = 999
        while maxbyes > self._maxbyes_allowed or num_rounds > self._maxrounds_allowed:
            maxbyes, num_rounds, games_scheduled  = self.scheduleGamesB()
    
        with open("valid_game_schedule.json", "w") as f:
            json.dump(games_scheduled, f, indent=4)

        return games_scheduled


    def loadGameSchedule(self, filename):
        with open(filename, "r") as f:
            games_scheduled = json.load(f)
        
        self._games = games_scheduled
        return


    def makeExcel(self):
        self.checkGameSchedule(self._games)
        from openpyxl import Workbook
        from openpyxl.styles import Font, Side, Border, Alignment
        wb = Workbook()
        ws = wb.active

        # styles definition
        bold_font = Font(bold=True)
        italic_font = Font(italic=True)
        center_alignment = Alignment(horizontal="center")
        double_side = Side(border_style="double")

        # List team numbers
        for i in range(self._num_pairs):
            # follows: zero + stride*block + offset
            cellname = chr(65 + 0*0 + 1)  + (30 + 1*i + 0).__str__()
            ws[cellname] = i+1
            ws[cellname].font = bold_font

        # List court numbers
        for i in range(self._num_courts):
            cellname = chr(65 + 5*i + 3)  + (1 + 0*0 + 0).__str__()
            ws[cellname] = "Court " + (i+1).__str__()
            ws[cellname].font = bold_font
            ws[cellname].alignment = center_alignment
 
        game_excel_index_list = []
        played_gamecounts = np.zeros(self._num_pairs, dtype=int)        
        for index_game,game in enumerate(self._games):
            r = index_game // self._num_courts
            c = index_game % self._num_courts
            cellname = chr(65) + (1 + 2*r + 1).__str__()
            ws[cellname] = "round " + (r+1).__str__()
            if len(game) > 0:
                cellname = chr(65 + 5*c + 3) + (1 + 2*r + 1).__str__()
                ws[cellname] = "vs"
                ws[cellname].font = italic_font
                ws[cellname].alignment = center_alignment
                cellname = chr(65 + 5*c + 3) + (1 + 2*r + 2).__str__()
                ws[cellname] = ":"
                ws[cellname].alignment = center_alignment
                # border
                cellname = chr(65 + 5*c + 1) + (1 + 2*r + 1).__str__()
                ws[cellname].border = Border(top=double_side, left=double_side)
                cellname = chr(65 + 5*c + 5) + (1 + 2*r + 1).__str__()
                ws[cellname].border = Border(top=double_side, right=double_side)
                cellname = chr(65 + 5*c + 1) + (1 + 2*r + 2).__str__()
                ws[cellname].border = Border(bottom=double_side, left=double_side)
                cellname = chr(65 + 5*c + 5) + (1 + 2*r + 2).__str__()
                ws[cellname].border = Border(bottom=double_side, right=double_side)
                for i in range(65 + 5*c + 2, 65 + 5*c + 5):
                    cellname = chr(i) + (1 + 2*r + 1).__str__()
                    ws[cellname].border = Border(top=double_side)
                    cellname = chr(i) + (1 + 2*r + 2).__str__()
                    ws[cellname].border = Border(bottom=double_side)



            for index_pair, pair in enumerate(game):
                # List pairs for a game
                if index_pair <= 1:
                    pair_offset = 1
                else:
                    pair_offset = 2
                cellname = chr(65 + 1*index_pair + (5*c+pair_offset)) + (1 + 2*r + 1).__str__()
                ws[cellname] = pair + 1
                ws[cellname].font = bold_font
                ws[cellname].alignment = center_alignment
           
                # score equation for a game 
                cellname = chr(65 + 1*played_gamecounts[pair] + 2)  + (30 + 1*pair + 0).__str__()
                if index_pair <= 1:
                    score_offset = 2
                else:
                    score_offset = 4
                ws[cellname] = "=" + chr(65 + 5*c + score_offset) + (1 + 2*r + 2).__str__()
                played_gamecounts[pair] += 1

        wb.save('schedule.xlsx')


def main():
    kob = KoB(num_pairs=32, num_courts=5, num_games_per_pair=8, maxbyes_allowed=2, maxrounds_allowed=13)

#    kob.makeGames()
#    kob.checkGames()
#    kob.saveGameSchedule()

    kob.loadGameSchedule("valid_game_schedule.32.json")
    kob.makeExcel()


if __name__=="__main__":
    main()
